"""
dashboard_builder.py
Generic "upload any data -> dashboard" engine.

Builds an .xlsx (or .xlsm, if a macro template is available) with:
  - Dashboard   : title (+ OMAC Developers capsule), KPI cards (formulas
                  reading cells on the Summary Data sheet), and one or two
                  charts per dimension.
  - Summary Data: per-dimension aggregate tables (category -> measure),
                  Pareto-sorted, which drive every chart and KPI card.
  - Data        : cleaned source as an Excel Table.
  - Config      : hidden sheet listing the chosen dimensions and measure
                  settings, read by the embedded VBA macro (if present).

NOTE: This version intentionally does NOT hand-write any native
PivotTable / PivotCache XML itself. Hand-built pivot cache parts almost
never satisfy Excel's schema validator, which causes the "we found a
problem with some content" repair prompt on open -- and Excel's
auto-repair then deletes the pivot parts AND the cell ranges they "own",
wiping out the literal fallback values the charts/KPIs depended on,
leaving a blank dashboard.

If MACRO_TEMPLATE_PATH points to an existing .xlsm containing a
Workbook_Open macro (see ThisWorkbook_macro.vba), build_workbook() starts
from that template with keep_vba=True so the macro is carried through
byte-for-byte. On open, that macro uses Excel's own PivotCaches/
PivotTables/Slicers APIs (reading the hidden Config sheet written here)
to build real, slicer-connected PivotTables and PivotCharts -- which
Excel always authors validly, since Excel is writing its own file.

If no template is present, build_workbook() falls back to the previous
plain .xlsx behaviour (static charts off the Summary Data sheet). Either
way the user can also select the 'SourceData' table on the Data sheet and
use Insert -> PivotTable themselves.
"""

import os
import re
import numpy as np
import pandas as pd
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.data_source import AxDataSource, StrRef, StrData, StrVal, NumRef, NumData, NumVal, NumDataSource
from openpyxl.drawing.text import CharacterProperties

NAVY = '1F4E5C'
TEAL = '2E8B8B'
LIGHT = 'EAF2F2'
ACCENT = 'C0392B'
GREY = '95A5A6'
GOLD = 'D4AC0D'
PALETTE = [TEAL, NAVY, ACCENT, 'E67E22', 'F1C40F', '27AE60', '8E44AD', '34495E']

RECORD_COL = '__Records__'
BLANK = '(Blank)'

# Path to the macro-enabled template (see ThisWorkbook_macro.vba for the
# one-time setup that creates this file). If present, build_workbook()
# returns a .xlsm with real, slicer-connected PivotTables/PivotCharts.
MACRO_TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'dashboard_macro_template.xlsm')

MAX_DIMENSIONS = 8
MAX_CHART_CATEGORIES = 10
DONUT_THRESHOLD = 5  # <= this many categories -> donut, else bar


def using_macro_template():
    """True if dashboard_macro_template.xlsm is present, in which case
    build_workbook() returns an .xlsm with real PivotTables/PivotCharts/
    Slicers (built by the embedded macro on open)."""
    return os.path.exists(MACRO_TEMPLATE_PATH)


# Name of the sheet that holds the per-dimension aggregate tables that
# drive every chart and KPI formula on the Dashboard.
SUMMARY_SHEET = 'Summary Data'

BRAND = 'OMAC Developers \u00b7 S M Baqir'


# ---------------------------------------------------------------------
# Step 1 - load & profile
# ---------------------------------------------------------------------
def load_dataframe(file_obj, filename):
    if filename.lower().endswith('.csv'):
        return pd.read_csv(file_obj)
    return pd.read_excel(file_obj)


def _norm_text(v):
    if not isinstance(v, str):
        return v
    v = v.replace('\n', ' ').replace('\r', ' ')
    v = ' '.join(v.split())
    return v


def clean_dataframe(df):
    """Clean headers AND normalize whitespace/newlines inside text cells.

    This is the generic fix for categories silently splitting into
    near-duplicates (e.g. "Finance\\n" vs "Finance " vs "Finance  Dept")
    which otherwise fragments groupings.
    """
    df = df.copy()
    cols = []
    seen = {}
    for c in df.columns:
        c2 = _norm_text(str(c))
        if not c2 or c2.lower().startswith('unnamed'):
            c2 = 'Column'
        if c2 in seen:
            seen[c2] += 1
            c2 = f'{c2} ({seen[c2]})'
        else:
            seen[c2] = 0
        cols.append(c2)
    df.columns = cols

    for col in df.columns:
        if df[col].dtype == object or pd.api.types.is_string_dtype(df[col]):
            df[col] = df[col].map(_norm_text)

    return df


def profile_columns(df):
    """Return dimension/measure candidates + a suggested default dimension set."""
    n = len(df)
    dim_candidates, num_candidates, scored = [], [], []

    id_pattern = r'\b(id|name|email|phone|contact|address|mrn|uid|url|date|time|no|number|code)\b'

    for col in df.columns:
        s = df[col]
        nunique = s.nunique(dropna=True)
        fill_rate = s.notna().sum() / n if n else 0
        looks_like_id_name = bool(re.search(id_pattern, col.lower()))
        looks_like_id_vals = nunique == n and n > 10

        if pd.api.types.is_numeric_dtype(s):
            if not looks_like_id_name and not looks_like_id_vals:
                num_candidates.append(col)

        is_long_text = False
        if s.dtype == object or pd.api.types.is_string_dtype(s):
            try:
                avg_len = s.dropna().astype(str).str.len().mean()
            except Exception:
                avg_len = 0
            is_long_text = bool(avg_len and avg_len > 30)

        is_numeric = pd.api.types.is_numeric_dtype(s)
        numeric_too_granular = is_numeric and nunique > 20

        looks_like_datetime_vals = False
        if s.dtype == object or pd.api.types.is_string_dtype(s):
            sample = s.dropna().astype(str)
            if len(sample):
                dt_re = re.compile(
                    r'\d{1,2}\s*(?:am|pm)|\d{1,2}[-/]\d{1,2}[-/]\d{2,4}|'
                    r'\d{1,2}[-\s](?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)', re.I)
                match_frac = sample.str.contains(dt_re).mean()
                looks_like_datetime_vals = match_frac > 0.3

        if (2 <= nunique <= max(50, int(n * 0.6)) and not is_long_text
                and not looks_like_id_vals and not looks_like_id_name
                and not numeric_too_granular and not looks_like_datetime_vals
                and fill_rate >= 0.3):
            dim_candidates.append(col)
            score = -nunique  # prefer fewer categories, but include long-tail too
            scored.append((score, col))

    scored.sort(reverse=True, key=lambda t: t[0])
    suggested = [c for _, c in scored[:6]]

    return {
        'dimension_candidates': dim_candidates,
        'numeric_candidates': num_candidates,
        'suggested_dimensions': suggested,
        'row_count': n,
    }


# ---------------------------------------------------------------------
# Step 2 - chart helpers
# ---------------------------------------------------------------------
def _style_title(chart, text):
    chart.title = text
    try:
        chart.title.tx.rich.p[0].r[0].rPr = CharacterProperties(sz=1200, b=True, solidFill=NAVY)
    except Exception:
        pass


def _text_categories(sheet_name, col_letter, min_row, max_row, values):
    """AxDataSource with strRef + cached strData, so charts render category
    labels correctly even without an Excel refresh (and so they're typed as
    text rather than openpyxl's default numRef, which is wrong for text)."""
    ref = f"'{sheet_name}'!${col_letter}${min_row}:${col_letter}${max_row}"
    cache = StrData(ptCount=len(values), pt=[StrVal(idx=i, v=str(v)) for i, v in enumerate(values)])
    return AxDataSource(strRef=StrRef(f=ref, strCache=cache))


def _num_values(sheet_name, col_letter, min_row, max_row, values):
    """NumDataSource (numRef + cached numCache) for series values."""
    ref = f"'{sheet_name}'!${col_letter}${min_row}:${col_letter}${max_row}"
    cache = NumData(ptCount=len(values), pt=[NumVal(idx=i, v=v) for i, v in enumerate(values)])
    return NumDataSource(numRef=NumRef(f=ref, numCache=cache))


# ---------------------------------------------------------------------
# Step 3 - build the workbook
# ---------------------------------------------------------------------
def build_workbook(df, dims, measure_mode, measure_col=None,
                    title='Auto-Generated Dashboard', source_label='',
                    chart_types=None):
    """
    df            : pandas DataFrame (already loaded)
    dims          : list of column names to use as breakdown/category columns (1-8)
    measure_mode  : 'count' | 'sum' | 'average'
    measure_col   : numeric column name, required if measure_mode != 'count'
    chart_types   : optional dict {dim_name: 'auto' | 'donut' | 'bar'}.
                     'auto' (default) picks donut for <= DONUT_THRESHOLD
                     categories and bar otherwise. Any dim not present in
                     the dict is treated as 'auto'.

    Returns: bytes of the .xlsx file
    """
    chart_types = chart_types or {}
    if not dims:
        raise ValueError('Select at least one dimension column.')
    dims = dims[:MAX_DIMENSIONS]

    df = clean_dataframe(df)
    df = df.reset_index(drop=True)
    n = len(df)
    df[RECORD_COL] = 1
    headers = list(df.columns)
    ncols = len(headers)

    for d in dims:
        df[d] = df[d].apply(lambda v: BLANK if pd.isna(v) else str(v))

    # -------- measure / aggregation setup --------
    if measure_mode == 'count':
        subtotal = 'count'
        measure_label = 'Records'
        numfmt = '#,##0'
        sort_fn = lambda d: df[d].value_counts()
    elif measure_mode == 'sum':
        if not measure_col:
            raise ValueError('measure_col required for sum')
        subtotal = 'sum'
        measure_label = f'Sum of {measure_col}'
        numfmt = '#,##0.00'
        sort_fn = lambda d: df.groupby(d)[measure_col].sum(min_count=1)
    else:
        if not measure_col:
            raise ValueError('measure_col required for average')
        subtotal = 'average'
        measure_label = f'Average {measure_col}'
        numfmt = '#,##0.00'
        sort_fn = lambda d: df.groupby(d)[measure_col].mean()

    # -------- categorical ordering (Pareto: highest value first) --------
    CAT_FIELDS = {}
    CAT_VALUES = {}
    GRAND_TOTAL = {}
    for d in dims:
        idx = headers.index(d)
        ser = sort_fn(d).fillna(0).sort_values(ascending=False)
        CAT_FIELDS[idx] = [str(v) for v in ser.index.tolist()]
        CAT_VALUES[idx] = [float(v) for v in ser.tolist()]
        if subtotal == 'sum':
            GRAND_TOTAL[idx] = float(df[measure_col].sum())
        elif subtotal == 'average':
            GRAND_TOTAL[idx] = float(df[measure_col].mean())
        else:
            GRAND_TOTAL[idx] = float(n)

    # -------- layout: where each dimension's summary table sits --------
    summary_layout = {}  # dim -> (header_row, grand_row, n_items, fld_idx, caption_row)
    running_row = 2
    for d in dims:
        idx = headers.index(d)
        n_items = len(CAT_FIELDS[idx])
        header_row = running_row + 1
        grand_row = header_row + 1 + n_items
        summary_layout[d] = (header_row, grand_row, n_items, idx, running_row)
        running_row = grand_row + 2

    # -------- workbook --------
    if os.path.exists(MACRO_TEMPLATE_PATH):
        wb = load_workbook(MACRO_TEMPLATE_PATH, keep_vba=True)
        # Repurpose the template's first sheet as 'Data' and drop any others
        # so we start from a clean slate (the macro rebuilds everything else
        # on open anyway).
        existing_sheets = list(wb.sheetnames)
        ws_data = wb[existing_sheets[0]]
        ws_data.title = 'Data'
        for name in existing_sheets[1:]:
            del wb[name]
        if ws_data.max_row > 0:
            ws_data.delete_rows(1, ws_data.max_row)
        for tbl_name in list(ws_data.tables.keys()):
            del ws_data.tables[tbl_name]
    else:
        wb = Workbook()
        ws_data = wb.active
        ws_data.title = 'Data'

    header_fill = PatternFill('solid', start_color=NAVY)
    header_font = Font(bold=True, color='FFFFFF', name='Calibri')
    for j, h in enumerate(headers, start=1):
        c = ws_data.cell(row=1, column=j, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(wrap_text=True, vertical='center')

    for i, (_, row) in enumerate(df.iterrows(), start=2):
        for j, h in enumerate(headers, start=1):
            val = row[h]
            if pd.isna(val):
                val = None
            elif isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val)
            cell = ws_data.cell(row=i, column=j, value=val)
            cell.font = Font(name='Calibri', size=10)
            cell.alignment = Alignment(vertical='top')

    last_col = get_column_letter(ncols)
    tbl = Table(displayName='SourceData', ref=f'A1:{last_col}{n + 1}')
    tbl.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False,
                                         showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws_data.add_table(tbl)
    ws_data.freeze_panes = 'A2'
    for j, h in enumerate(headers, start=1):
        ws_data.column_dimensions[get_column_letter(j)].width = 22 if len(h) > 14 else 14
    rc_idx = headers.index(RECORD_COL) + 1
    ws_data.column_dimensions[get_column_letter(rc_idx)].hidden = True

    # ---- Summary Data sheet (drives charts + KPI cards) ----
    ws_sum = wb.create_sheet(SUMMARY_SHEET)
    ws_sum.sheet_view.showGridLines = False
    note = ws_sum.cell(row=1, column=1)
    note.value = ('Summary tables behind every chart and KPI on the Dashboard. '
                   'Want to slice the raw data yourself? Go to the Data sheet, '
                   'click inside the table, then Insert \u2192 PivotTable.')
    note.font = Font(name='Calibri', size=11, italic=True, color=NAVY)
    note.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws_sum.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws_sum.row_dimensions[1].height = 36
    for cell in ws_sum['A1:L1'][0]:
        cell.fill = PatternFill('solid', start_color=LIGHT)

    for d in dims:
        h, g, n_items, idx, caption_row = summary_layout[d]
        cap = ws_sum.cell(row=caption_row, column=1)
        cap.value = d
        cap.font = Font(name='Calibri', size=12, bold=True, color=NAVY)

        hdr1 = ws_sum.cell(row=h, column=1, value=d)
        hdr2 = ws_sum.cell(row=h, column=2, value=measure_label)
        for cell in (hdr1, hdr2):
            cell.font = Font(name='Calibri', bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', start_color=NAVY)
        for k in range(n_items):
            lc = ws_sum.cell(row=h + 1 + k, column=1, value=CAT_FIELDS[idx][k])
            vc = ws_sum.cell(row=h + 1 + k, column=2, value=CAT_VALUES[idx][k])
            vc.number_format = numfmt
        glc = ws_sum.cell(row=g, column=1, value='Grand Total')
        gvc = ws_sum.cell(row=g, column=2, value=GRAND_TOTAL[idx])
        glc.font = Font(bold=True)
        gvc.font = Font(bold=True)
        gvc.number_format = numfmt

    ws_sum.column_dimensions['A'].width = 28
    ws_sum.column_dimensions['B'].width = 16

    # ---- Dashboard sheet ----
    ws_dash = wb.create_sheet('Dashboard', 0)
    _build_dashboard(ws_dash, ws_sum, dims, headers, CAT_FIELDS, CAT_VALUES, summary_layout,
                      measure_mode, measure_col, measure_label, numfmt, title, source_label, n,
                      chart_types)

    # ---- hidden Config sheet (read by the ThisWorkbook_macro.vba on open) ----
    ws_cfg = wb.create_sheet('Config')
    ws_cfg.cell(row=1, column=1, value=len(dims))
    ws_cfg.cell(row=1, column=2, value=measure_mode)
    ws_cfg.cell(row=2, column=2, value=measure_col or '')
    ws_cfg.cell(row=3, column=2, value=measure_label)
    for i, d in enumerate(dims, start=1):
        ws_cfg.cell(row=i + 1, column=1, value=d)
    ws_cfg.sheet_state = 'hidden'

    wb._sheets = [wb['Dashboard'], wb[SUMMARY_SHEET], wb['Data'], wb['Config']]
    wb.active = 0

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


# ---------------------------------------------------------------------
# Step 4 - Dashboard sheet (branding + KPI cards + charts)
# ---------------------------------------------------------------------
def _build_dashboard(ws, ws_sum, dims, headers, CAT_FIELDS, CAT_VALUES, summary_layout,
                       measure_mode, measure_col, measure_label, numfmt, title, source_label, n,
                       chart_types=None):
    chart_types = chart_types or {}
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    n_dims = len(dims)
    n_cols_layout = max(24, (min(n_dims, 6) * 3) or 18)
    last_col_letter = get_column_letter(n_cols_layout)

    # ---------------- Title band ----------------
    ws.merge_cells(f'A1:{get_column_letter(n_cols_layout - 6)}2')
    c = ws['A1']
    c.value = title
    c.font = Font(name='Calibri', size=22, bold=True, color='FFFFFF')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    for row in ws[f'A1:{last_col_letter}2']:
        for cell in row:
            cell.fill = PatternFill('solid', start_color=NAVY)

    # OMAC branding capsule, top-right of the title band
    badge_c0 = n_cols_layout - 5
    ws.merge_cells(start_row=1, start_column=badge_c0, end_row=1, end_column=n_cols_layout)
    b = ws.cell(row=1, column=badge_c0)
    b.value = BRAND
    b.font = Font(name='Calibri', size=10, bold=True, color=NAVY)
    b.fill = PatternFill('solid', start_color=GOLD)
    b.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=2, start_column=badge_c0, end_row=2, end_column=n_cols_layout)
    b2 = ws.cell(row=2, column=badge_c0)
    b2.value = 'Auto-Generated Dashboard'
    b2.font = Font(name='Calibri', size=8, italic=True, color='FFFFFF')
    b2.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(f'A3:{last_col_letter}3')
    c = ws['A3']
    subtitle = f'{n:,} records'
    if source_label:
        subtitle = f'{source_label}  |  {subtitle}'
    c.value = subtitle
    c.font = Font(name='Calibri', size=11, italic=True, color=GREY)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    # ---------------- KPI cards ----------------
    if measure_mode == 'count':
        kpi1_label = 'Total Records'
    elif measure_mode == 'sum':
        kpi1_label = f'Total {measure_col}'
    else:
        kpi1_label = f'Overall Average {measure_col}'

    d0 = dims[0]
    h0, g0, n0, idx0, _ = summary_layout[d0]
    SD = SUMMARY_SHEET
    grand_addr0 = f"'{SD}'!B{g0}"
    top_label_addr0 = f"'{SD}'!A{h0 + 1}"
    top_val_addr0 = f"'{SD}'!B{h0 + 1}"
    second_label_addr0 = f"'{SD}'!A{h0 + 2}" if n0 >= 2 else top_label_addr0
    second_val_addr0 = f"'{SD}'!B{h0 + 2}" if n0 >= 2 else top_val_addr0
    cat_count_formula = f"COUNTA('{SD}'!A{h0 + 1}:A{h0 + n0})"

    kpis = [
        (kpi1_label, f'={grand_addr0}', numfmt, NAVY),
        (f'Top {d0}', f'={top_label_addr0}', '@', TEAL, True),
        (f'{measure_label} ({d0} top)', f'={top_val_addr0}', numfmt, TEAL),
    ]

    if len(dims) >= 2:
        d1 = dims[1]
        h1, g1, n1, idx1, _ = summary_layout[d1]
        kpis.append((f'Top {d1}', f"='{SD}'!A{h1 + 1}", '@', ACCENT, True))
        kpis.append((f'{measure_label} ({d1} top)', f"='{SD}'!B{h1 + 1}", numfmt, ACCENT))
    else:
        kpis.append((f'2nd Highest {d0}', f'={second_label_addr0}', '@', ACCENT, True))
        kpis.append((f'{measure_label} (2nd)', f'={second_val_addr0}', numfmt, ACCENT))

    kpis.append((f'{d0} Categories', f'={cat_count_formula}', '#,##0', GREY))

    card_width = 3
    start_row = 5
    end_row = 8
    thin = Side(style='thin', color='D0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for idx, kpi in enumerate(kpis[:6]):
        label, formula, fmt, color = kpi[0], kpi[1], kpi[2], kpi[3]
        is_text = len(kpi) > 4 and kpi[4]
        c0 = idx * card_width + 1
        c1 = c0 + card_width - 1
        rng_label = f'{get_column_letter(c0)}{start_row}:{get_column_letter(c1)}{start_row}'
        rng_value = f'{get_column_letter(c0)}{start_row + 1}:{get_column_letter(c1)}{end_row - 1}'
        ws.merge_cells(rng_label)
        ws.merge_cells(rng_value)
        lab = ws.cell(row=start_row, column=c0, value=label.upper())
        lab.font = Font(name='Calibri', size=9, bold=True, color=GREY)
        lab.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        val = ws.cell(row=start_row + 1, column=c0, value=formula)
        val.font = Font(name='Calibri', size=20 if is_text else 24, bold=True, color=color)
        if fmt != '@':
            val.number_format = fmt
        val.alignment = Alignment(horizontal='center', vertical='center', wrap_text=is_text)
        for r in range(start_row, end_row + 1):
            for col in range(c0, c1 + 1):
                cell = ws.cell(row=r, column=col)
                cell.border = border
                cell.fill = PatternFill('solid', start_color=LIGHT if r == start_row else 'FFFFFF')

    # ---------------- Charts: top-N chart, plus an "all categories" chart underneath ----------------
    anchor_cols_idx = [1, 14]       # column A and N
    ROW_CM = 0.529  # approx cm per default row height

    row_cursor = 10
    row_group_height = 0
    for i, d in enumerate(dims):
        h, g, n_items, idx, _ = summary_layout[d]
        col_idx = i % 2
        if col_idx == 0 and i > 0:
            row_cursor += row_group_height + 2
            row_group_height = 0
        c0 = anchor_cols_idx[col_idx]
        anchor = f'{get_column_letter(c0)}{row_cursor}'

        choice = chart_types.get(d, 'auto')
        if choice == 'donut':
            use_donut = True
        elif choice == 'bar':
            use_donut = False
        else:
            use_donut = n_items <= DONUT_THRESHOLD

        top_n = min(n_items, MAX_CHART_CATEGORIES)

        if use_donut:
            chart = DoughnutChart()
            chart.add_data(Reference(ws_sum, min_col=2, min_row=h, max_row=h + top_n), titles_from_data=True)
            chart.height = 8
            chart.width = 9.5
            chart.dataLabels = DataLabelList(showVal=True, showCatName=False, showSerName=False,
                                              showPercent=False, showLegendKey=False)
            chart.holeSize = 55
            title_text = f'{d}  (top {top_n} of {n_items})' if n_items > top_n else d
            _style_title(chart, title_text)
            chart.series[0].cat = _text_categories(ws_sum.title, 'A', h + 1, h + top_n, CAT_FIELDS[idx][:top_n])
            chart.series[0].val = _num_values(ws_sum.title, 'B', h + 1, h + top_n, CAT_VALUES[idx][:top_n])
            chart.series[0].dPt = [
                DataPoint(idx=k, spPr=GraphicalProperties(solidFill=PALETTE[k % len(PALETTE)]))
                for k in range(top_n)
            ]
            chart_height_rows = 16
        else:
            chart = BarChart()
            chart.type = 'bar'
            chart.add_data(Reference(ws_sum, min_col=2, min_row=h + 1, max_row=h + top_n), titles_from_data=False)
            chart.height = 9.5
            chart.width = 12.5
            chart.dataLabels = DataLabelList(showVal=True, showCatName=False, showSerName=False,
                                              showPercent=False, showLegendKey=False)
            chart.series[0].cat = _text_categories(ws_sum.title, 'A', h + 1, h + top_n, CAT_FIELDS[idx][:top_n])
            chart.series[0].val = _num_values(ws_sum.title, 'B', h + 1, h + top_n, CAT_VALUES[idx][:top_n])
            chart.series[0].graphicalProperties = GraphicalProperties(solidFill=PALETTE[i % len(PALETTE)])
            chart.x_axis.axPos = 'l'
            chart.y_axis.axPos = 'b'
            chart.x_axis.delete = False
            chart.y_axis.delete = False
            chart.x_axis.tickLblPos = 'nextTo'
            chart.y_axis.tickLblPos = 'nextTo'
            chart.x_axis.crosses = 'autoZero'
            chart.y_axis.crosses = 'autoZero'
            chart.y_axis.majorGridlines = None
            chart.legend = None
            title_text = f'{d}  (top {top_n} of {n_items})' if n_items > top_n else d
            _style_title(chart, title_text)
            chart_height_rows = 19

        ws.add_chart(chart, anchor)
        pair_height = chart_height_rows

        # ---- second chart: ALL categories, directly underneath ----
        if n_items > top_n:
            all_height_cm = max(9.5, n_items * 0.35 + 2)
            all_height_rows = int(all_height_cm / ROW_CM) + 1
            anchor2_row = row_cursor + chart_height_rows + 1
            anchor2 = f'{get_column_letter(c0)}{anchor2_row}'

            chart_all = BarChart()
            chart_all.type = 'bar'
            chart_all.add_data(Reference(ws_sum, min_col=2, min_row=h + 1, max_row=h + n_items), titles_from_data=False)
            chart_all.height = all_height_cm
            chart_all.width = 12.5
            chart_all.dataLabels = DataLabelList(showVal=True, showCatName=False, showSerName=False,
                                                  showPercent=False, showLegendKey=False)
            chart_all.series[0].cat = _text_categories(ws_sum.title, 'A', h + 1, h + n_items, CAT_FIELDS[idx][:n_items])
            chart_all.series[0].val = _num_values(ws_sum.title, 'B', h + 1, h + n_items, CAT_VALUES[idx][:n_items])
            chart_all.series[0].graphicalProperties = GraphicalProperties(solidFill=PALETTE[(i + 1) % len(PALETTE)])
            chart_all.x_axis.axPos = 'l'
            chart_all.y_axis.axPos = 'b'
            chart_all.x_axis.delete = False
            chart_all.y_axis.delete = False
            chart_all.x_axis.tickLblPos = 'nextTo'
            chart_all.y_axis.tickLblPos = 'nextTo'
            chart_all.x_axis.crosses = 'autoZero'
            chart_all.y_axis.crosses = 'autoZero'
            chart_all.y_axis.majorGridlines = None
            chart_all.legend = None
            _style_title(chart_all, f'{d}  (all {n_items} categories)')

            ws.add_chart(chart_all, anchor2)
            pair_height += 1 + all_height_rows

        row_group_height = max(row_group_height, pair_height)

    # ---------------- Column widths / row heights ----------------
    for col in range(1, n_cols_layout + 1):
        ws.column_dimensions[get_column_letter(col)].width = 8.5
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 18
    for r in range(5, 9):
        ws.row_dimensions[r].height = 22
